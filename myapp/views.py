# -*- coding: utf-8 -*-
"""
Created on Sat Aug  3 21:04:15 2024

@author: elisa
"""

from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib import messages
from django.conf import settings
from openpyxl import load_workbook, Workbook
import os


excel_path = os.path.join(settings.BASE_DIR, 'Résultat_Test.xlsx')

images_path = 'static/images/'
font_style = "Broadway, 26"
font_color = "Purple"

def check_create_excel_file(path):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Nom-Prénom", "Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7", "Q8", "Q9"])
        wb.save(path)
        print(f"Created new Excel file: {path}")

def add_to_excel(path, data):
    print(f"Adding to Excel file: {data}")
    wb = load_workbook(path)
    ws = wb.active
    ws.append(data)
    wb.save(path)
    print(f"Added to Excel file successfully: {data}")

def update_excel(path, row, col):
    print(f"Updating Excel file at row {row}, column {col}")
    wb = load_workbook(path)
    ws = wb.active
    if ws.cell(row=row, column=col).value is None:
        ws.cell(row=row, column=col).value = 1
    else:
        ws.cell(row=row, column=col).value += 1
    wb.save(path)
    print(f"Updated Excel file successfully at row {row}, column {col}")

def index(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        if name and name.replace(' ', '').isalnum():
            add_to_excel(excel_path, [name])
            row = load_workbook(excel_path).active.max_row
            print(f"New row for name {name}: {row}")
            return redirect(reverse('image', kwargs={'image_num': 3, 'row': row}))
        else:
            messages.error(request, "Saisie invalide. Veuillez saisir votre prénom et nom.")
    return render(request, 'myapp/index.html')

def image(request, image_num, row):
    if image_num > 10:
        return redirect(reverse('thank_you'))

    if request.method == 'POST':
        response = request.POST.get('response')
        if response.isdigit() and 1 <= int(response) <= 9:
            update_excel(excel_path, row, int(response) + 1)
            return redirect(reverse('image', kwargs={'image_num': image_num + 1, 'row': row}))
        else:
            messages.error(request, "Saisie invalide. Veuillez saisir un chiffre de 1 à 9.")

    prompts = [
        "Votre ville préférée est : ",
        "Quel serait votre tailleur pantalon préféré : ",
        "Kilos superflus ? Votre réaction ? ",
        "Quelle chambre choisiriez-vous ? ",
        "Quelle est votre réaction si on vous complimente sur votre style ? ",
        "Dans quel pyjama dormirez-vous le mieux ? ",
        "Quel genre de coiffure adopteriez-vous ? ",
        "Que reflète votre image dans le miroir ? "
    ]
    prompt = prompts[image_num-3] if image_num-3 < len(prompts) else ""

    return render(request, 'myapp/image.html', {'image_num': image_num, 'prompt': prompt})

def thank_you(request):
    return render(request, 'myapp/thank_you.html')
